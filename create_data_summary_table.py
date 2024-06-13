# -*- coding: utf-8 -*-
"""
Created on Thu Jun 13 14:38:19 2024

@author: ludoa
"""
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# Step 1: Convert TSV to Excel
tsv_file_path = "E:/BIDS_dataset_longitudinale/dataset_v2/participants.tsv"
excel_file_path = "E:/BIDS_dataset_longitudinale/dataset_v2/data_summary.xlsx"

# Read the TSV file
df = pd.read_csv(tsv_file_path, sep='\t')

# Define the BIDS directory root
bids_root = "E:/BIDS_dataset_longitudinale/dataset_v2"  # Replace this with your actual BIDS root directory

# Initialize columns for validation data
sessions = ['v1', 'v2', 'v3']


for ses in sessions:
    df[f'{ses}_valid_fmri'] = False
    df[f'{ses}_valid_dmri'] = False

# Function to validate subject-session pairs
def validate_subject_sessions(bids_root):
    validation_data = {}
    
    for sub in os.listdir(bids_root):
        if sub.startswith('sub-'):
            sub_path = os.path.join(bids_root, sub)
            sub_data = {}
            for ses_ID in os.listdir(sub_path):            
                if ses_ID.startswith('ses-'):
                    ses = ses_ID.replace("ses-","")
                    ses_path = os.path.join(sub_path, ses_ID)
                    anat_path = os.path.join(ses_path, 'anat')
                    func_path = os.path.join(ses_path, 'func')
                    fmap_path = os.path.join(ses_path, 'fmap')
                    dwi_path = os.path.join(ses_path, 'dwi')

                    
                    anat_exists = os.path.exists(anat_path)
                    func_exists = os.path.exists(func_path)
                    fmap_exists = os.path.exists(fmap_path)
                    dwi_exists = os.path.exists(dwi_path)

                    valid_fmri = anat_exists and func_exists and fmap_exists
                    valid_dmri = anat_exists and dwi_exists
                    
                    sub_data[ses] = {
                        'valid_fmri': valid_fmri,
                        'valid_dmri': valid_dmri
                    }
            
            validation_data[sub] = sub_data
    
    return validation_data

# Get validation data
validation_data = validate_subject_sessions(bids_root)

# Update the DataFrame with validation data
for idx, row in df.iterrows():
    subject = row['participant_id']
    if subject in validation_data:
        for ses in sessions:
            if ses in validation_data[subject]:
                df.at[idx, f'{ses}_valid_fmri'] = validation_data[subject][ses]['valid_fmri']
                df.at[idx, f'{ses}_valid_dmri'] = validation_data[subject][ses]['valid_dmri']

# Save updated DataFrame to Excel
df.to_excel(excel_file_path, index=False)
print(f"Excel file updated at: {excel_file_path}")


# Load the workbook and access the sheet
wb = load_workbook(excel_file_path)
ws = wb.active

# Define the color fill styles
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

# Apply conditional formatting
for col in ws.iter_cols(min_col=len(df.columns)-len(sessions)*2+1, max_col=len(df.columns), min_row=2, max_row=ws.max_row):
    for cell in col:
        cell_value = cell.value
        if cell_value is True:
            cell.fill = green_fill
        elif cell_value is False:
            cell.fill = red_fill
            

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
wb.save(excel_file_path)
